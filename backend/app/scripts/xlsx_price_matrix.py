"""Parses the "материал × поставщик" price-matrix xlsx (updated supplier
price list, replaces the earlier two-CSV format — materials.csv had a
derived, not sourced, `internal_sku`; this file has none at all, see
`_next_sku` below) into the same (materials, prices) shape
`import_real_data.py` already knows how to load.

One sheet, one row per material, one column per supplier price. Real data in
this file only occupies the first ~300 rows despite the sheet nominally
having 1090 — the rest is blank formatting. `Group` (category) is
forward-filled in the source: set once on the first row of a category block,
blank on every row after until the next block starts — NOT one row per
category, has to be carried forward while parsing, blanks are not missing
data. A handful of rows are section dividers (no supplier prices at all,
`Quantity == "-"`, a garbled/mojibake description) and must be dropped, not
imported as materials — see `_is_divider_row`.

Structure was inspected directly (openpyxl, read_only) before writing this,
not assumed from the old CSV shape — see the conversation this was written
in for the full breakdown of row/column counts.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

HEADER_ROW = 1
FIRST_DATA_ROW = 2

COL_GROUP = 1
COL_DESCRIPTION = 2
COL_QUANTITY = 3

# Sheet header text (col index -> our supplier name) -> SUPPLIER_NAMES in
# import_real_data.py. Header cells are the actual literal strings found in
# row 1 of the source file, including the stray leading newline on
# "\nLancing" and the abbreviated/typo'd names — normalized (stripped) before
# lookup, not before storing this map.
SUPPLIER_COLUMN_HEADERS: dict[str, str] = {
    "EMS": "EMS",
    "Lancing": "Lancing",
    "Fasteners": "JM Fasteners",
    "Auminum D": "Aluminum Distributors Int LLC",
    "Florida Sales": "Florida Sales & Marketing",
    "AMS": "American Metals Supply",
    "Classic Metals": "Classic Metals",
}

# Same prefix scheme as the old materials.csv (DOOR-001, DOOR-002, ... —
# see git history / docs/DEPLOYMENT.md context) so generated SKUs stay in a
# recognizable style. Not guaranteed to line up 1:1 with the previous
# internal_sku values — this file's row order/count differs slightly, and
# the old prefixes were themselves derived, not a business identifier taken
# from any source system.
CATEGORY_SKU_PREFIX: dict[str, str] = {
    "Doors": "DOOR",
    "Gutter": "GUTR",
    "Mesh": "MESH",
    "Connectors": "CONN",
    "Profil": "PROF",
    "Roof panels": "ROOF",
    "Screws": "SCRW",
    "Caulk": "CAUL",
}
FALLBACK_SKU_PREFIX = "MISC"

FALLBACK_UNIT = "pcs"
"""Applied when Quantity is blank — matches the overwhelming majority
(198/294 real rows) of the sheet's own values. Flagged per-row in
ParsedWorkbook.rows_with_fallback_unit so the dry-run summary surfaces it
for manual review rather than silently blending into the pcs count."""


def _is_divider_row(description: str, quantity_raw: object) -> bool:
    """Section-divider rows carry no material data: garbled description,
    `Quantity == "-"`, zero prices in any supplier column. Matched on
    Quantity alone (cheap, exact) rather than trying to detect "garbled
    text", which has no reliable signal of its own — every divider row in
    the source happens to use literal "-", which no real material row
    uses (verified: no other row in the file has Quantity == "-")."""
    return str(quantity_raw).strip() == "-"


def _clean_unit(quantity_raw: object) -> tuple[str, bool]:
    """Returns (unit, used_fallback). The source's Quantity column is a
    free-text descriptor ("1 pcs", "1 box/100 pcs", "compl"), not a clean
    unit code — Material.unit is stored as-is (String(20)) rather than
    parsed further, since there is no reliable rule to reduce e.g. "1
    box/100 pcs" to a single token without losing information the business
    put there on purpose."""
    if quantity_raw is None or str(quantity_raw).strip() == "":
        return FALLBACK_UNIT, True
    return str(quantity_raw).strip(), False


def _clean_price(raw: object) -> float | None:
    """Source has a handful of comma-decimal cells (e.g. "3,5" instead of
    3.5) mixed in with numeric cells — European/Russian decimal notation
    typo'd into an otherwise US-formatted sheet. Only strings get the
    comma->dot treatment; a plain numeric cell is returned as-is."""
    if raw is None or str(raw).strip() == "":
        return None
    if isinstance(raw, int | float):
        return float(raw)
    text = str(raw).strip().replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _next_sku(category: str | None, counters: dict[str, int]) -> str:
    prefix = CATEGORY_SKU_PREFIX.get(category or "", FALLBACK_SKU_PREFIX)
    counters[prefix] = counters.get(prefix, 0) + 1
    return f"{prefix}-{counters[prefix]:03d}"


@dataclass
class MaterialRow:
    internal_sku: str
    description: str
    category: str | None
    unit: str
    used_fallback_unit: bool
    row_number: int  # 1-indexed sheet row, for diagnostics


@dataclass
class PriceCell:
    description: str
    supplier_name: str
    price: float


@dataclass
class ParsedWorkbook:
    materials: list[MaterialRow] = field(default_factory=list)
    prices: list[PriceCell] = field(default_factory=list)
    divider_rows_skipped: int = 0
    rows_with_fallback_unit: list[str] = field(default_factory=list)  # descriptions
    unparseable_price_cells: list[tuple[int, str, object]] = field(default_factory=list)
    unmapped_supplier_headers: list[str] = field(default_factory=list)


def parse_price_matrix(path: Path) -> ParsedWorkbook:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header = next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))
    supplier_columns: dict[int, str] = {}
    result = ParsedWorkbook()
    for col_idx, cell in enumerate(header):
        if col_idx in (COL_GROUP, COL_DESCRIPTION, COL_QUANTITY) or cell is None:
            continue
        header_text = str(cell).strip()
        if header_text == "":
            continue
        mapped = SUPPLIER_COLUMN_HEADERS.get(header_text)
        if mapped is None:
            result.unmapped_supplier_headers.append(header_text)
            continue
        supplier_columns[col_idx] = mapped

    sku_counters: dict[str, int] = {}
    current_category: str | None = None

    for row_number, row in enumerate(
        ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True), start=FIRST_DATA_ROW
    ):
        description = row[COL_DESCRIPTION]
        if description is None or str(description).strip() == "":
            continue  # trailing blank-formatting rows past the real data
        description = str(description).strip()

        quantity_raw = row[COL_QUANTITY]

        group_cell = row[COL_GROUP]
        if group_cell is not None and str(group_cell).strip() != "":
            current_category = str(group_cell).strip()

        if _is_divider_row(description, quantity_raw):
            result.divider_rows_skipped += 1
            continue

        unit, used_fallback = _clean_unit(quantity_raw)
        if used_fallback:
            result.rows_with_fallback_unit.append(description)

        sku = _next_sku(current_category, sku_counters)
        result.materials.append(
            MaterialRow(
                internal_sku=sku,
                description=description,
                category=current_category,
                unit=unit,
                used_fallback_unit=used_fallback,
                row_number=row_number,
            )
        )

        for col_idx, supplier_name in supplier_columns.items():
            raw_price = row[col_idx] if col_idx < len(row) else None
            if raw_price is None or str(raw_price).strip() == "":
                continue
            price = _clean_price(raw_price)
            if price is None:
                result.unparseable_price_cells.append((row_number, supplier_name, raw_price))
                continue
            result.prices.append(
                PriceCell(description=description, supplier_name=supplier_name, price=price)
            )

    return result
